"""
Author: Mike Fuson
Purpose: Created to help my brother quickly determine 
    which crew members are working in Nashville EMS.

Notes: Simplifies scheduling and reduces time spent
    figuring out shift assignments.

PS: Love You Bro
"""

import pandas as pd
import re
import math
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import textwrap
import tkinter as tk
from tkinter import filedialog
import os
import sys
from copy import copy
from pathlib import Path

ON_DUTY_CODES = [
    "EDIEMSEV15EDI",
    "EMSSHFDIF",
    "EMSSHFDIFES09",
    "EMSSHFDIFES33",
    "STWEP",
    "STWEA",
    "*STWEP",
    "*STWEA",
    "+OTEMS15",
    "+OOCEDTCPM",
    "+OOCEDTCAM",
    "EDI15-SE",
    "+OOCEDTCAMES21",
    "+OOCEDTCPMES21",
    "EDIEMSEV15EDI",    
    "+EDIEMSEV15EDI",
    "+OTEMS15-SS",
    "OTS15",
    "EMSSHFDIFEMSS01",
    "+OOCAC"
]
GENERIC_ON_DUTY_CODES = [
    "."
]
NOT_WORK_CODES = [
    "*ESTNWPM",
    "*ESTNWAM",
    ".ESTNWAM",
    ".ESTNWPM",
    ".HPM",
    ".HAM",
    ".PFLMPMMaternity",
    ".PFLMAMMaternity",
    ".FMLA-AM",
    ".FMLA-PM",
    ".PLPM",
    ".PLAM",
    ".EMS MLAM3",
    ".EMS MLPM3",
    ".PLPM",
    ".PLAM",
    ".VAM",
    ".LWOP",
    ".FMLA-AMChild",
    "+S120PM0.",
    "+S120PM0.",
    ".SASC",
    ".EVPM",
    ".FMLA-AMChild",
    ".VPM"
]

# Mapping of old field names to new ones a result of the excel import
FIELD_RENAME_MAP = {
    'Unnamed: 2': 'ID',
    'Unnamed: 3': 'Name',
    'Unnamed: 5': 'Code',
    'Unnamed: 6': 'From',
    'Unnamed: 7': 'Through',
    'Unnamed: 8': 'Hours'
}

# Apply renaming to each record before cleaning
def rename_fields(records):
    renamed_records = []
    for entry in records:
        renamed_entry = {}
        for k, v in entry.items():
            new_key = FIELD_RENAME_MAP.get(k, k)  # Replace if in map, else keep original
            renamed_entry[new_key] = v
        renamed_records.append(renamed_entry)
    return renamed_records

def get_report_dates(rosterFileName):
    wb = load_workbook(rosterFileName, data_only=True)
    ws = wb.active  # or wb["Sheet1"]
    centerHeader = ws.oddHeader.center.text if ws.oddHeader else None
    centerFooter = ws.oddFooter.center.text if ws.oddFooter else None

    m = re.search(r"\[(\d{2}/\d{2}/\d{4})\]", centerHeader)
    if m:
        report_date = m.group(1)
        report_date = report_date.replace("/", "-")
    else:
        print("No Report Date Found")
        report_date = ''

    m = re.search(r"(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})", centerFooter)

    if m:
        generated_date = m.group(1)
        generated_date = generated_date.replace("/", "-")
    else:
        print("No Generated Date/Time Found")
        generated_date = ''

    return report_date,generated_date

def load_roster_report(rosterFileName, base_dir=None):
    # If a base_dir is given and the filename is not already absolute
    if base_dir and not os.path.isabs(rosterFileName):
        file_path = os.path.join(base_dir, rosterFileName)
    else:
        file_path = rosterFileName

    # Normalize to absolute path
    file_path = os.path.abspath(file_path)

    df = pd.read_excel(file_path)
    data = df.to_dict(orient='records')
    cleanData = clean_fields(data)
    readableFieldsData=rename_fields(cleanData)
    
    return readableFieldsData

def is_empty_or_nan(value):
    return (
        value is None or
        (isinstance(value, float) and math.isnan(value)) or
        (isinstance(value, str) and value.strip() == '')
    )

def clean_fields(records):
    cleaned = []
    for entry in records:
        new_entry = {}

        for k, v in entry.items():
            # Skip keys with empty values
            if is_empty_or_nan(v):
                continue

            # Remove 'Unnamed: X' keys with empty values
            if re.match(r'^Unnamed: \d+$', k):
                if not is_empty_or_nan(v):
                    new_entry[k] = v
            else:
                new_entry[k] = v

        # Skip entries that are completely empty OR contain 'Rank' as any value
        if new_entry and 'Rank' not in [str(v).strip() for v in new_entry.values()]:
            cleaned.append(new_entry)

    return cleaned

def propagate_group(data, label_key="Group"):
    result = []
    current_label = None

    for entry in data:
        if isinstance(entry, dict):
            keys = list(entry.keys())

            # If it's a single-key dict (label)
            if len(keys) == 1:
                current_label = entry[keys[0]]
                continue  # Skip adding this entry to result

            # Apply the current label to the entry
            if current_label is not None:
                entry[label_key] = current_label

            result.append(entry)

    return result

def filter_on_duty(data, allowed_codes, cancel_codes, generic_codes):
    """
    Filters a list of dicts (with "Name" and "Code") to find who is on duty.

    On duty means:
      - A person has at least one allowed code (direct match or contains a generic code)
      - And they do not have any cancel codes.

    Args:
        data (list[dict]): Each dict has "Name" and "Code".
        allowed_codes (set[str]): Exact allowed code strings.
        cancel_codes (set[str]): Exact cancel code strings.
        generic_codes (set[str]): Generic substrings that may appear inside a Code.

    Returns:
        list[dict]: Filtered rows for on-duty names, only with allowed/generic codes.
    """

    def is_allowed(code):
        # Exact allowed OR contains any generic substring
        if code in allowed_codes:
            return True
        return any(g in code for g in generic_codes)

    # Build Name -> set of codes
    codes_by_name = {}
    for entry in data:
        name = entry.get("Name")
        code = entry.get("Code")
        if not name or not code:
            continue
        codes_by_name.setdefault(name, set()).add(code)

    # Determine who is on duty
    on_duty_names = set()
    for name, codes in codes_by_name.items():
        has_allowed = any(is_allowed(c) for c in codes)
        has_cancel = any(c in cancel_codes for c in codes)
        if has_allowed and not has_cancel:
            on_duty_names.add(name)

    # Return filtered rows
    result = []
    for entry in data:
        name = entry.get("Name")
        code = entry.get("Code")
        if name in on_duty_names and is_allowed(code):
            result.append(entry)

    return result

def assign_shift(data):
    for entry in data:
        from_time = entry.get("From")
        through_time = entry.get("Through")

        if from_time == "06:00": # and through_time == "18:00":
            entry["Shift"] = "AM"
        elif from_time == "18:00": # and through_time == "06:00":
            entry["Shift"] = "PM"
        else:
            entry["Shift"] = "Special"
    return data

def find_unique_lables(data, label_key="group"):
    unique_groups = set()
    labels=[]
    for entry in data:
        if isinstance(entry, dict) and label_key in entry:
            unique_groups.add(entry[label_key])

    for group in sorted(unique_groups):
        labels.append(group)
    return labels

def truncate_group_values(data, separator="/"):
    """
    Truncate the 'Group' field of each dictionary in a list to everything after the separator.

    Args:
        data (list of dict): Input array of dictionaries.
        separator (str): The delimiter to split on (default is "/").

    Returns:
        list of dict: Updated list with truncated 'Group' values.
    """
    for item in data:
        if "Group" in item and isinstance(item["Group"], str):
            # Split and strip spaces around the separator
            parts = item["Group"].split(separator)
            if len(parts) > 1:
                item["Group"] = parts[-1].strip()
    return data

def sort_by_second_string(data):
    """
    Sorts a list of lists by the first string in each sublist.

    Args:
        data (list): A list of lists where the first element of each sublist is a string.

    Returns:
        list: A new list sorted by the first element of each sublist.
    """
    return sorted(data, key=lambda x: x[1])


# -------- Styles --------
thin = Side(border_style="thin", color="000000")
row_border = Border(top=thin, left=thin, right=thin, bottom=thin)

fill_white = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
fill_gray  = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")  # light gray

medium = Side(border_style="medium", color="000000")  # darker/thicker border
header_border = Border(top=thin, left=thin, right=thin, bottom=thin)

header_fill = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")  # gray background

def append_header(ws, values, min_col=1, max_col=None, exclude_cols=None):
    """
    Append a header row:
      - Thick black borders
      - Bold text
      - Font size +1 pt from default
      - Optional gray shading
    """
    if exclude_cols is None:
        exclude_cols = []

    ws.append(values)
    r = ws.max_row

    if max_col is None:
        max_col = len(values)

    for c in range(min_col, max_col + 1):
        if c in exclude_cols:
            continue
        cell = ws.cell(row=r, column=c)
        # Apply thick border
        cell.border = header_border
        # Bold and +1 pt font
        base_size = cell.font.size or 11
        cell.font = Font(name=cell.font.name, size=base_size + 1, bold=True)
        # Shading
        cell.fill = fill_white
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Enable wrap text if cell has newline
        if cell.value is not None and "\n" in str(cell.value):
            cell.alignment = Alignment(wrapText=True,horizontal="center", vertical="center")

# -------- Append helper --------
def append_row_with_format(ws, values, min_col=1, max_col=None, header_rows=1,exclude_cols=None):
    """
    Append a row of values, apply borders and alternating fill,
    skipping formatting for specified columns.

    exclude_cols: list of column indexes (1-based) to skip, e.g. [2, 4]
    """
    if exclude_cols is None:
        exclude_cols = []

    ws.append(values)
    r = ws.max_row
    if max_col is None:
        max_col = len(values)

    # Decide fill color (alternate starting after header_rows)
    band_idx = r - header_rows
    if band_idx <= 0:
        row_fill = fill_white
    else:
        row_fill = fill_gray if band_idx % 2 == 0 else fill_white

    # Apply style only if not excluded
    for c in range(min_col, max_col + 1):
        if c in exclude_cols:
            continue
        cell = ws.cell(row=r, column=c)
        cell.border = row_border
        cell.fill = row_fill
        cell.alignment = Alignment(wrapText=True)
        # # Enable wrap text if cell has newline
        # if cell.value is not None and "\n" in str(cell.value):
        #     cell.alignment = Alignment(wrapText=True)

def autofit_row_heights(ws, col_widths_chars):
    """
    Roughly estimate needed row heights for wrapped text.
    Works on already-filled rows in ws.
    """
    line_height = 15.0  # points per line (approx for Calibri 11pt)
    padding = 2.0       # breathing room

    for r in range(1, ws.max_row + 1):
        max_lines = 1
        for c in range(1, min(ws.max_column, len(col_widths_chars)) + 1):
            cell = ws.cell(row=r, column=c)
            text = "" if cell.value is None else str(cell.value)
            width_chars = max(1, int(col_widths_chars[c - 1]))

            para_lines = 0
            for para in text.splitlines() or [""]:
                wrapped = textwrap.wrap(para, width=width_chars) or [""]
                para_lines += len(wrapped)

            max_lines = max(max_lines, para_lines)

        ws.row_dimensions[r].height = max_lines * line_height + padding


def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # e.g. "A", "B", "C"
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2  # add some padding
        ws.column_dimensions[col_letter].width = adjusted_width
def append_first_sheet_with_formatting(src_path, dest_path, new_sheet_name=None):
    """
    Copy the first worksheet (with formatting) from src_path into dest_path as a new sheet.
    Returns the name of the newly created sheet.
    """
    src_path = Path(src_path)
    dest_path = Path(dest_path)

    if not src_path.exists():
        raise FileNotFoundError(f"Source not found: {src_path}")
    if not dest_path.exists():
        raise FileNotFoundError(f"Destination not found: {dest_path}")

    keep_vba_src = src_path.suffix.lower() == ".xlsm"
    keep_vba_dest = dest_path.suffix.lower() == ".xlsm"

    src_wb = load_workbook(src_path, data_only=False, keep_vba=keep_vba_src)
    src_ws = src_wb.worksheets[0]  # first sheet

    dest_wb = load_workbook(dest_path, data_only=False, keep_vba=keep_vba_dest)

    # Resolve target name without collisions
    base_name = new_sheet_name or (src_ws.title or "Sheet1")
    name = base_name
    i = 1
    while name in dest_wb.sheetnames:
        i += 1
        name = f"{base_name} ({i})"

    dest_ws = dest_wb.create_sheet(title=name)

    # --- Cells: values + styles (guarded) ---
    for row in src_ws.iter_rows():
        for cell in row:
            tgt = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            # Copy number format and core style parts if present
            try:
                tgt.number_format = cell.number_format
            except Exception:
                pass
            if getattr(cell, "has_style", False):
                try:
                    if getattr(cell, "font", None):
                        tgt.font = copy(cell.font)
                    if getattr(cell, "fill", None):
                        tgt.fill = copy(cell.fill)
                    if getattr(cell, "border", None):
                        tgt.border = copy(cell.border)
                    if getattr(cell, "alignment", None):
                        tgt.alignment = copy(cell.alignment)
                    if getattr(cell, "protection", None):
                        tgt.protection = copy(cell.protection)
                except Exception:
                    # Ignore any exotic style that openpyxl can't clone
                    pass

    # --- Column dimensions ---
    try:
        for col_letter, dim in src_ws.column_dimensions.items():
            td = dest_ws.column_dimensions[col_letter]
            if dim.width is not None:
                td.width = dim.width
            if getattr(dim, "hidden", None) is not None:
                td.hidden = dim.hidden
            if getattr(dim, "outlineLevel", None):
                td.outlineLevel = dim.outlineLevel
            if getattr(dim, "bestFit", None):
                td.bestFit = dim.bestFit
    except Exception:
        pass

    # --- Row dimensions ---
    try:
        for idx, dim in src_ws.row_dimensions.items():
            td = dest_ws.row_dimensions[idx]
            if dim.height is not None:
                td.height = dim.height
            if getattr(dim, "hidden", None) is not None:
                td.hidden = dim.hidden
            if getattr(dim, "outlineLevel", None):
                td.outlineLevel = dim.outlineLevel
    except Exception:
        pass

    # --- Merged cells ---
    try:
        for merged_range in getattr(src_ws.merged_cells, "ranges", []):
            dest_ws.merge_cells(str(merged_range))
    except Exception:
        pass

    # --- Freeze panes ---
    try:
        dest_ws.freeze_panes = src_ws.freeze_panes
    except Exception:
        pass

    # --- Sheet tab color (this is where your error came from) ---
    # Access the attribute directly; guard the worksheet objects themselves.
    try:
        if src_ws is not None and dest_ws is not None:
            tab_color = getattr(getattr(src_ws, "sheet_properties", None), "tabColor", None)
            if tab_color is not None:
                # Ensure dest has sheet_properties
                if getattr(dest_ws, "sheet_properties", None) is not None:
                    dest_ws.sheet_properties.tabColor = copy(tab_color)
    except Exception:
        pass

    # --- Print settings ---
    try:
        if getattr(src_ws, "print_area", None):
            dest_ws.print_area = src_ws.print_area

        # Page setup
        try:
            dest_ws.page_setup.orientation = src_ws.page_setup.orientation
            dest_ws.page_setup.paperSize = src_ws.page_setup.paperSize
            dest_ws.page_setup.fitToPage = src_ws.page_setup.fitToPage
            dest_ws.page_setup.fitToHeight = src_ws.page_setup.fitToHeight
            dest_ws.page_setup.fitToWidth = src_ws.page_setup.fitToWidth
            dest_ws.page_setup.scale = src_ws.page_setup.scale
        except Exception:
            pass

        # Margins
        try:
            dm, sm = dest_ws.page_margins, src_ws.page_margins
            dm.left = sm.left; dm.right = sm.right
            dm.top = sm.top; dm.bottom = sm.bottom
            dm.header = sm.header; dm.footer = sm.footer
        except Exception:
            pass

        # Print options / titles
        try:
            dest_ws.print_title_cols = src_ws.print_title_cols
            dest_ws.print_title_rows = src_ws.print_title_rows
        except Exception:
            pass

        try:
            dpo, spo = dest_ws.print_options, src_ws.print_options
            dpo.horizontalCentered = spo.horizontalCentered
            dpo.verticalCentered = spo.verticalCentered
            dpo.gridLines = spo.gridLines
        except Exception:
            pass
    except Exception:
        pass

    # --- Data validations ---
    try:
        dvs = getattr(src_ws, "data_validations", None)
        if dvs and getattr(dvs, "count", 0) > 0:
            dest_ws.data_validations = copy(dvs)
    except Exception:
        pass

    dest_wb.save(dest_path)

def create_ouput_spreadsheet(data,outFileName,reportDate,generatedDateTime):

    #Create workbook and worksheet
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "ON Duty"
    ws2 = wb.create_sheet(title="Travelers")
    ws3 = wb.create_sheet(title="EDU-REACH-ETC")

    #Collect all unique column names (keys)
    medicHeader = ['Shift','Unit','Paramedic','EMT','3rd Employee','Shift','Unit','Paramedic','EMT','3rd Employee']
    cheifHeader = ['Shift','Unit','Chief','','','Shift','Unit','Chief']
    OtherHeader = ['','Times', 'Unit', 'Name']
    #Truncated the Group String
    data = truncate_group_values(data, separator="/")

    #Write each row of data
    groupLabels = find_unique_lables(data,'Group')
    groupLabelsSorted = sorted(groupLabels, key=lambda x: x.split("/", 1)[-1])


    asstRow = []
    medicRows =[]
    distRows = []
    specialRows =[]
    travelAMRows = []    
    travelPMRows = []
    AsstFound = False

    for group in groupLabelsSorted:
        entries = [entry for entry in data if group in entry.get('Group', '')]
        if ('Medic' in group)and not('73' in group):
            AMentries = [AMentry for AMentry in entries if 'AM' in AMentry.get('Shift', '')]
            PMentries = [PMentry for PMentry in entries if 'PM' in PMentry.get('Shift', '')]           

            if len(AMentries) > 3:
                print("More than 3 Employees on {} AM".format(group))
                AMrow = [AMentries[0]['Shift'],AMentries[0]['Group'],AMentries[0]['Name'],AMentries[1]['Name'],AMentries[2]['Name']]  
            elif len(AMentries) == 3:
                AMrow = [AMentries[0]['Shift'],AMentries[0]['Group'],AMentries[0]['Name'],AMentries[1]['Name'],AMentries[2]['Name']]
            elif len(AMentries) == 2:
                AMrow = [AMentries[0]['Shift'],AMentries[0]['Group'],AMentries[0]['Name'],AMentries[1]['Name'],'']
            elif len(AMentries) == 1:
                print("Less than 2 Employees on {} AM".format(group))
                AMrow = [AMentries[0]['Shift'],AMentries[0]['Group'],AMentries[0]['Name'],'','']
            elif len(AMentries) == 0:                
                print("No Employees on {} AM".format(group))
                AMrow = ['','','','','']

            if len(PMentries) > 3:
                print("More than 3 Employees on {} PM".format(group))
                PMrow = [PMentries[0]['Shift'],PMentries[0]['Group'],PMentries[0]['Name'],PMentries[1]['Name'],PMentries[2]['Name']]  
            elif len(PMentries) == 3:
                PMrow = [PMentries[0]['Shift'],PMentries[0]['Group'],PMentries[0]['Name'],PMentries[1]['Name'],PMentries[2]['Name']]
            elif len(PMentries) == 2:
                PMrow = [PMentries[0]['Shift'],PMentries[0]['Group'],PMentries[0]['Name'],PMentries[1]['Name'],'']
            elif len(PMentries) == 1:
                print("Less than 2 Employees on {} PM".format(group))
                PMrow = [PMentries[0]['Shift'],PMentries[0]['Group'],PMentries[0]['Name'],'','']
            elif len(PMentries) == 0:                              
                print("No Employees on {} PM".format(group))
                PMrow = ['','','','','']

            if not(all(item == '' for item in (AMrow+PMrow))):
                medicRows.append(AMrow+PMrow)

        elif 'On-Duty' in group:
            AsstFound = True
            AMentries = [AMentry for AMentry in entries if 'AM' in AMentry.get('Shift', '')]
            PMentries = [PMentry for PMentry in entries if 'PM' in PMentry.get('Shift', '')]

            if len(AMentries) == 1:
                AMASSrow = [AMentries[0]['Shift'],'ASST',AMentries[0]['Name']]
            elif len(AMentries) == 0:                              
                print("No ASS on Duty {} PM".format(group))
                AMASSrow = ['AM','ASST','']

            if len(PMentries) == 1:
                PMASSrow = [PMentries[0]['Shift'],'ASST',PMentries[0]['Name']]
            elif len(PMentries) == 0:                              
                print("No ASS on Duty {} PM".format(group))
                PMASSrow = ['PM','ASST','']

            if not(all(item == '' for item in (AMASSrow+PMASSrow))):
                asstRow = (AMASSrow+['','']+PMASSrow)
                
        elif ('District' in group) and not('EMS' in group):   
            AMentries = [AMentry for AMentry in entries if 'AM' in AMentry.get('Shift', '')]
            PMentries = [PMentry for PMentry in entries if 'PM' in PMentry.get('Shift', '')]

            if len(AMentries) == 1:
                AMdistrow = [AMentries[0]['Shift'],AMentries[0]['Group'],AMentries[0]['Name']]
            elif len(AMentries) == 0:                              
                print("No Chief on {} PM".format(group))
                AMdistrow = ['','','']

            if len(PMentries) == 1:
                PMdistrow = [PMentries[0]['Shift'],PMentries[0]['Group'],PMentries[0]['Name']]
            elif len(PMentries) == 0:                              
                print("No Chief on {} PM".format(group))
                PMdistrow = ['','','']

            if not(all(item == '' for item in (AMdistrow+PMdistrow))):
                distRows.append(AMdistrow+['','']+PMdistrow)
        elif 'Travelers AM' in group:
            if len(entries) != 0:
                for travNum in range(len(entries)):
                    travelAMRow = ['',entries[travNum]['From']+'-'+entries[travNum]['Through'],entries[travNum]['Group'],entries[travNum]['Name']]
                    if not(all(item == '' for item in (travelAMRow))):
                        travelAMRows.append(travelAMRow)
        elif 'Travelers PM' in group:
            if len(entries) != 0:
                for travNum in range(len(entries)):
                    travelPMRow = ['',entries[travNum]['From']+'-'+entries[travNum]['Through'],entries[travNum]['Group'],entries[travNum]['Name']]
                    if not(all(item == '' for item in (travelPMRow))):
                        travelPMRows.append(travelPMRow)            
        else:            
            if len(entries) != 0:
                for specNum in range(len(entries)):
                    specialRow = ['',entries[specNum]['From']+'-'+entries[specNum]['Through'],entries[specNum]['Group'],entries[specNum]['Name']]
                    if not(all(item == '' for item in (specialRow))):
                        specialRows.append(specialRow)

    if not(AsstFound):
        asstRow = ['AM','ASST','','','','PM','ASST','']
        print("No ASSes on Duty")
    ReportDateRow        = ['','','On-Duty Roster Report\n'+reportDate]
    GeneratedDateTimeRow = ['','','Roster Report Generated\n'+generatedDateTime]
    DoubleReportDateRow        = ['','','On-Duty Roster Report\n'+reportDate,'','','','','On-Duty Roster Report\n'+reportDate]
    DoubleGeneratedDateTimeRow = ['','','Roster Report Generated\n'+generatedDateTime,'','','','','Roster Report Generated\n'+generatedDateTime]

    specialRows = sort_by_second_string(specialRows)
    append_header(ws1,DoubleReportDateRow,exclude_cols=[1,2,4,5,6,7])
    append_header(ws1,DoubleGeneratedDateTimeRow,exclude_cols=[1,2,4,5,6,7])
    ws1.append([])
    append_header(ws1,cheifHeader,exclude_cols=[4,5])
    chiefHead_row = ws1.max_row
    append_row_with_format(ws1,asstRow,header_rows=chiefHead_row,exclude_cols=[4,5])
    for row in distRows:
        append_row_with_format(ws1,row,header_rows=chiefHead_row,exclude_cols=[4,5])
    ws1.append([])
    append_header(ws1,medicHeader)
    medicHead_row = ws1.max_row
    for row in medicRows:
        append_row_with_format(ws1,row,header_rows=medicHead_row)
    ws1.append([])

    append_header(ws2, ReportDateRow,max_col=4, exclude_cols=[1, 2, 4, 5, 6, 7])
    append_header(ws2, GeneratedDateTimeRow,max_col=4, exclude_cols=[1, 2, 4, 5, 6, 7])
    append_header(ws2,OtherHeader,exclude_cols=[1])
    for row in travelAMRows:
        append_row_with_format(ws2,row,exclude_cols=[1])
    ws2.append([])
    for row in travelPMRows:
        append_row_with_format(ws2,row,exclude_cols=[1])

    append_header(ws3, ReportDateRow,max_col=4, exclude_cols=[1, 2, 4, 5, 6, 7])
    append_header(ws3, GeneratedDateTimeRow,max_col=4, exclude_cols=[1, 2, 4, 5, 6, 7])
    append_header(ws3, OtherHeader, exclude_cols=[1])
    ws3.append([])

    for row in specialRows:
        append_row_with_format(ws3,row,exclude_cols=[1])

    col_widths = [10, 45, 45, 45, 45, 10, 45, 45, 45, 45]  # adjust as needed

    # Alignment for wrapping
    wrap_align = Alignment(wrap_text=True, vertical="top")

    # --- 1) Apply fixed widths to ws1 ---
    for idx, w in enumerate(col_widths, start=1):
        ws1.column_dimensions[get_column_letter(idx)].width = w

    # --- 2) Apply wrapping to all existing cells ---
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row,
                             min_col=1, max_col=len(col_widths)):
        for cell in row:
            cell.alignment = wrap_align

    autofit_row_heights(ws1, col_widths)
    ws1.page_setup.pageOrder = "overThenDown"
    ws1.page_setup.orientation = "landscape"
    ws1.page_setup.scale = None
    ws1.sheet_properties.pageSetUpPr.fitToPage = True
    ws1.page_setup.fitToWidth = 2
    ws1.page_setup.fitToHeight = 1  # 0 means "as many as needed"
    ws1.page_margins.left = 0.80  # default is 0.7
    ws1.page_margins.right = 0.80  # default is 0.7
    ws1.page_margins.top = 0.50  # default is 0.75
    ws1.page_margins.bottom = 0.50  # default is 0.75
    ws1.page_margins.header = 0.3  # default is 0.3
    ws1.page_margins.footer = 0.3  # default is 0.3

    auto_adjust_column_width(ws2)
    ws2.page_setup.pageOrder = "overThenDown"
    ws2.page_setup.orientation = "landscape"
    ws2.page_setup.scale = None
    ws2.sheet_properties.pageSetUpPr.fitToPage = True
    ws2.page_setup.fitToWidth = 1
    ws2.page_setup.fitToHeight = 0  # 0 means "as many as needed"

    auto_adjust_column_width(ws3)
    ws3.page_setup.pageOrder = "overThenDown"
    ws3.page_setup.orientation = "landscape"
    ws3.page_setup.scale = None
    ws3.sheet_properties.pageSetUpPr.fitToPage = True
    ws3.page_setup.fitToWidth = 1
    ws3.page_setup.fitToHeight = 0  # 0 means "as many as needed"

    wb.save(outFileName)

def browse_for_excel():
    # Create a hidden root window (so only the dialog shows)
    root = tk.Tk()
    root.withdraw()

    # Open file dialog
    file_paths = filedialog.askopenfilenames(
        title="Select Roster File(s)",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )

    root.destroy()
    if not file_paths:  # user canceled
        print("No input file selected. Exiting...")
        sys.exit(0)

    return list(file_paths)

def get_output_filename(reportDate,initialdir="."):
    import os, sys
    import tkinter as tk
    from tkinter import filedialog
    from datetime import datetime

    root = tk.Tk()
    root.withdraw()

    now_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
    default_name = f"On_Duty_Roster_{reportDate}.xlsx"

    # One popup: choose folder AND filename together
    path = filedialog.asksaveasfilename(
        title="Save As",
        initialdir=initialdir,
        initialfile=default_name,
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )

    root.destroy()

    if not path:
        print("No output file selected. Exiting...")
        sys.exit(0)

    # Ensure .xlsx extension (asksaveasfilename should add it, but double-check)
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"

    return os.path.abspath(path)

def main():
    inFileNames = browse_for_excel()
    for  inFileName in inFileNames:
        try:
            print("Input File Path: {} ".format(inFileName))  
            reportDate , generatedDateTime = get_report_dates(inFileName)
            roster = load_roster_report(inFileName)
            rosterWitGroup = propagate_group(roster)
            onDutyRoster = filter_on_duty(rosterWitGroup,ON_DUTY_CODES,NOT_WORK_CODES,GENERIC_ON_DUTY_CODES)
            onDutyShifts = assign_shift(onDutyRoster)
            outFileName = get_output_filename(reportDate)
            print("Output File Path: {} ".format(outFileName))
            create_ouput_spreadsheet(onDutyShifts,outFileName,reportDate,generatedDateTime)
            append_first_sheet_with_formatting(inFileName,outFileName,"Roster Report {}".format(reportDate))
        except Exception as e:
            print(f"Error processing {inFileName}: {e}")
            continue  #loop naturally continues anyway

if __name__ == "__main__":
    main()